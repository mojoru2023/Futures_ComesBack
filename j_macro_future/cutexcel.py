import openpyxl
from datetime import datetime
from openpyxl.styles import NamedStyle
from tqdm import tqdm  # 导入 tqdm 库以显示进度条


def cutexcel(filename, sheetname):
    # 加载 Excel 文件
    input_file = filename
    wb = openpyxl.load_workbook(input_file)

    # 获取指定名称的工作表
    ws = wb[sheetname]  # 指定读取名称为指定的工作表

    # 创建一个字典来存储按月份分组的数据
    data_by_month = {}

    # 获取总行数以便于显示进度条（减去标题行）
    total_rows = ws.max_row - 1

    # 读取数据（假设日期在第一列，从第二行开始）
    for row in tqdm(ws.iter_rows(min_row=2, values_only=True), total=total_rows, desc="处理数据"):
        date_value = row[0]  # 假设日期在第一列

        # 检查 date_value 是否为 None
        if date_value is None:
            continue  # 如果为空，跳过此行

        if isinstance(date_value, str):  # 如果是字符串，则需要转换
            try:
                date_value = datetime.strptime(date_value, '%m/%d/%Y')
            except ValueError:
                print(f"无效日期格式: {date_value}")
                continue
        elif isinstance(date_value, (int, float)):  # 如果是数字，可能是Excel序列号
            date_value = openpyxl.utils.datetime.from_excel(date_value)

        month = date_value.month
        if month not in data_by_month:
            data_by_month[month] = []

        data_by_month[month].append((date_value, *row[1:]))  # 保留日期并附加其他列内容

    # 定义日期格式样式
    date_style = NamedStyle(name='datetime', number_format='YYYY-MM-DD')

    # 将每个月的数据写入新的 Excel 文件
    for month, rows in data_by_month.items():
        new_workbook = openpyxl.Workbook()
        new_sheet = new_workbook.active

        # 写入标题行
        header = [cell.value for cell in ws[1]]  # 从原始工作表获取标题行
        new_sheet.append(header)  # 向新工作表添加标题行

        # 写入数据行
        for row in tqdm(rows, desc=f'写入月份 {month} 的数据'):
            new_row = list(row)  # 转换为列表
            new_sheet.append(new_row)  # 将数据添加到新工作表

            # 设置日期格式
            new_sheet.cell(row=new_sheet.max_row, column=1).style = date_style  # 日期在第一列

        # 保存新文件
        new_filename = f'{filename}-{month}.xlsx'
        new_workbook.save(new_filename)
        print(f"已保存：{new_filename}")

    print("所有月份的文件已切割并保存！")


if __name__ == "__main__":
    cutexcel("N225minif_2021.xlsx", "3min")
